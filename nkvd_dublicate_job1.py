import openpyxl

file_xls = '2009.xlsx'
wb = openpyxl.load_workbook(file_xls)

list_all_data = []

for page in ('1', '2'):
     wb_s = wb[page]
     for row in range(1, wb_s.max_row+1):
         list_all_data.append('_'.join([wb_s.cell(row=row, column=1).value,
                                        wb_s.cell(row=row, column=2).value]))

list_all_data = list(set(list_all_data))

wb_s = wb['3']
for persona in list_all_data:
    wb_s.append(persona.split('_'))

wb.save(file_xls)
wb.close()
